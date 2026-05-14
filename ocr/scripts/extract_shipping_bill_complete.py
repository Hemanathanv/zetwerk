#!/usr/bin/env python3
"""
Complete Shipping Bill extraction (text-layer PDF + ICEGATE QR) in Python.

1. Load final field schema from Final_Schema/shipping_bill_fields.xlsx (or fallback xlsx).
2. Extract text + tables per page with pdfplumber (text-based PDFs).
3. Decode CBIC QR via scripts/decode-icegate-sb-qr helpers (decimal -> 0x1b segments).
4. Heuristic text matching: label-style "FieldName:" lines in extracted text.
5. Emit one JSON with schema rows, QR payload, page stats, and per-field status.

Limitations:
- Scanned/image-only PDFs have little or no text layer; use OCR separately.
- Field matching is heuristic; tune patterns for your exact SB layout.
- QR binary tail is not verified cryptographically (use ICETRAK for that).

Usage:
  pip install -r requirements-python.txt
  python scripts/extract_shipping_bill_complete.py --pdf "Input/Shipping Bill/foo.pdf"
  python scripts/extract_shipping_bill_complete.py --pdf foo.pdf --schema "Final_Schema/shipping_bill_fields.xlsx" --out out.json
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import importlib.util


def _load_qr_module():
    path = ROOT / "scripts" / "decode-icegate-sb-qr.py"
    spec = importlib.util.spec_from_file_location("decode_icegate_sb_qr", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader
    spec.loader.exec_module(mod)
    return mod


_qr = _load_qr_module()
decode_from_decimal = _qr.decode_from_decimal
try_read_pdf_qr = _qr.try_read_pdf_qr


def resolve_schema_xlsx() -> Path:
    primary = ROOT / "Final_Schema" / "shipping_bill_fields.xlsx"
    if primary.exists():
        return primary
    fallback = ROOT / "output" / "schema-discovery" / "Shipping Bill.xlsx"
    if fallback.exists():
        primary.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(fallback, primary)
        return primary
    raise FileNotFoundError(
        f"Missing schema: {primary}. Add shipping_bill_fields.xlsx or run schema discovery "
        f"so {fallback} exists."
    )


def load_schema_rows(xlsx_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    hmap = {h.lower().replace(" ", ""): i for i, h in enumerate(header) if h}

    def col(*names: str) -> int | None:
        for n in names:
            k = n.lower().replace(" ", "")
            if k in hmap:
                return hmap[k]
        return None

    i_sec = col("section")
    i_fn = col("fieldname", "field name")
    i_ft = col("fieldtype", "field type")
    i_desc = col("description")
    i_req = col("required")
    if i_fn is None:
        raise ValueError("Schema xlsx must have FieldName or 'Field Name' column")

    out: list[dict] = []
    for row in rows[1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        fn = row[i_fn] if i_fn < len(row) else None
        if fn is None or str(fn).strip() == "":
            continue
        rec = {
            "section": str(row[i_sec]).strip() if i_sec is not None and i_sec < len(row) and row[i_sec] else "",
            "fieldName": str(fn).strip(),
            "fieldType": str(row[i_ft]).strip() if i_ft is not None and i_ft < len(row) and row[i_ft] else "string",
            "description": str(row[i_desc]).strip() if i_desc is not None and i_desc < len(row) and row[i_desc] else "",
            "required": False,
        }
        if i_req is not None and i_req < len(row) and row[i_req] is not None:
            v = row[i_req]
            rec["required"] = str(v).strip().lower() in ("yes", "true", "1", "y")
        out.append(rec)
    return out


def extract_pdf_plumber(pdf_path: Path) -> dict:
    import pdfplumber

    pages_out: list[dict] = []
    full_chunks: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            tables = page.extract_tables() or []
            pages_out.append(
                {
                    "page": i + 1,
                    "width": page.width,
                    "height": page.height,
                    "text": text,
                    "textCharCount": len(text),
                    "lineCount": len(text.splitlines()),
                    "tables": tables,
                    "tableCount": len(tables),
                }
            )
            full_chunks.append(text)
    full_text = "\n\n".join(full_chunks)
    return {
        "pageCount": len(pages_out),
        "pages": pages_out,
        "fullText": full_text,
        "fullTextCharCount": len(full_text),
    }


def qr_to_flat_dict(qr_decoded: dict) -> dict:
    """Map ASCII QR segments to a small flat dict for merging."""
    flat: dict[str, str] = {}
    for seg in qr_decoded.get("segments", []):
        if seg.get("encoding") != "ascii":
            continue
        hint = seg.get("hint", f"segment_{seg['index']}")
        flat[hint] = seg.get("value", "")
    # Convenience composite
    body = flat.get("sb_reference_body", "")
    if body:
        flat["printed_sb_number_guess"] = "SB22" + body
    return flat


def _label_regex_fragment(label: str) -> str:
    """Avoid 'ASSESS' matching inside 'Assessment'; allow multi-word labels."""
    t = label.strip()
    if not t:
        return re.escape(t)
    parts = t.split()
    if len(parts) == 1 and len(parts[0]) <= 8 and parts[0].isalpha() and parts[0].isupper():
        esc = re.escape(parts[0])
        return rf"(?<![A-Za-z]){esc}(?![A-Za-z])"
    return re.escape(t)


def clean_extracted_value(field_name: str, raw: str | None) -> str | None:
    """
    Remove echoed field labels and ICEGATE 'next column' tails (e.g. '... 8. GSTIN / TYPE').
    PDF text often has rows like '1.MODE 2.ASSESS 3.EXMN ...' — the capture after MODE must not
    include sibling headers.
    """
    if raw is None:
        return None
    v = raw.strip()
    if not v:
        return None

    fl = field_name.strip()
    if fl:
        # Drop leading "FieldName:", "FIELD NAME.", etc.
        v = re.sub(rf"(?is)^{re.escape(fl)}\s*[:./]?\s*", "", v, count=1).strip()
        v = re.sub(rf"(?is)^{re.escape(fl)}\s+", "", v, count=1).strip()

    # Sibling numbered field starts the "value" → wrong match (e.g. MODE → "2.ASSESS ...")
    if re.match(r"^\d{1,2}\.\s*[A-Za-z]", v):
        return None

    # Cut at next ICEGATE-style numbered field on the same logical line: " 15.PORT", " 8. GSTIN"
    cut = re.search(r"\s+(?=\d{1,2}\.\s*[A-Za-z0-9/])", v)
    if cut:
        v = v[: cut.start()].strip()

    # Cut at " digit(s). " when followed by space (e.g. "6390220 8. ")
    cut2 = re.search(r"\s+(?=\d{1,2}\.\s)", v)
    if cut2:
        v = v[: cut2.start()].strip()

    v = v.strip(" \t:;,-")
    if not v or v.lower() == fl.lower():
        return None

    # Still a ladder of numbered headers (many "N.LABEL" tokens)
    if len(v) > 40 and len(re.findall(r"\d{1,2}\.[A-Z]{2,}", v)) >= 2:
        return None

    # Short flag fields (MODE, LUT, …) should not be long prose
    short_tokens = (
        "MODE",
        "ASSESS",
        "EXMN",
        "JOBBING",
        "MEIS",
        "DBK",
        "RODTP",
        "LUT",
        "DFRC",
        "RE-EXP",
        "LICENCE",
    )
    if fl.upper() in short_tokens and len(v) > 20:
        return None

    return v


def find_value_in_text(field_name: str, section: str, full_text: str) -> tuple[str | None, str | None, float]:
    """
    Return (value, evidence_snippet, confidence).
    """
    if not full_text.strip():
        return None, None, 0.0
    ft = full_text.replace("\r", "")
    # Normalize field label variants
    candidates = [field_name.strip()]
    if "(" in field_name:
        candidates.append(field_name.split("(")[0].strip())

    best = (None, None, 0.0)
    for label in candidates:
        if len(label) < 2:
            continue
        frag = _label_regex_fragment(label)
        # Prefer same-line capture ending before next " N.LABEL" (numbered sibling)
        sibling = r"(?:\s+(?=\d{1,2}\.\s*[A-Za-z0-9/])|$)"
        for pat in (
            rf"(?im)^{frag}\s*[:.]?\s*(.+?){sibling}",
            rf"(?im){frag}\s*[:.]?\s*(.+?){sibling}(?:\n|$)",
        ):
            m = re.search(pat, ft)
            if m:
                val = m.group(1).strip()
                val = clean_extracted_value(field_name, val)
                if val:
                    return val, m.group(0)[:300], 0.75
        for pat in (
            rf"(?im)^{frag}\s*[:.]?\s*(.+)$",
            rf"(?im){frag}\s*[:.]?\s*(.+?)(?:\n|$)",
        ):
            m = re.search(pat, ft)
            if m:
                val = clean_extracted_value(field_name, m.group(1).strip())
                if val:
                    if len(val) > 200:
                        val = val[:200] + "…"
                    return val, m.group(0)[:300], 0.7
        # Loose: anchored label on same line (no substring matches inside longer words)
        for line in ft.splitlines():
            m = re.search(rf"(?i){frag}\s*[:.]?\s*(.*)", line)
            if not m:
                continue
            tail = m.group(1).strip()
            if tail.startswith(":"):
                tail = tail[1:].strip()
            tail = clean_extracted_value(field_name, tail)
            if tail and len(tail) < 500:
                conf = 0.45 if len(tail) > 3 else 0.25
                if conf > best[2]:
                    best = (tail, line[:250], conf)
    return best


def _rel(p: Path, root: Path) -> str:
    try:
        return str(p.resolve().relative_to(root.resolve()))
    except ValueError:
        return str(p.resolve())


def qr_hint_matches_field(field_name: str, section: str, qr_flat: dict) -> tuple[str | None, float]:
    """Very loose link from schema field name to QR hint keys."""
    fn = field_name.lower()
    sec = section.lower()
    pairs = [
        ("port", "port_code", 0.95),
        ("fob", "fob_value", 0.9),
        ("iec", "iec_suffix_or_fragment", 0.85),
        ("drawback", "drawback_or_duty_field", 0.85),
        ("package", "packages_or_count", 0.85),
        ("challan", "challan_or_cin", 0.85),
        ("cin", "challan_or_cin", 0.85),
        ("sb date", "sb_date", 0.9),
        ("shipping bill date", "sb_date", 0.9),
        ("sb no", "sb_number_fragment", 0.8),
        ("sb number", "sb_number_fragment", 0.8),
        ("shipping bill no", "sb_number_fragment", 0.8),
    ]
    for needle, key, conf in pairs:
        if needle in fn or needle in sec:
            v = qr_flat.get(key)
            if v:
                return str(v), conf
    if ("sb_reference" in fn) or ("reference" in fn and "sb" in fn):
        v = qr_flat.get("printed_sb_number_guess") or qr_flat.get("sb_reference_body")
        if v:
            return str(v), 0.85
    return None, 0.0


def run_extraction(pdf_path: Path, schema_path: Path) -> dict:
    schema = load_schema_rows(schema_path)
    pdf_text = extract_pdf_plumber(pdf_path)
    qr_decimal = try_read_pdf_qr(str(pdf_path))
    qr_decoded = decode_from_decimal(qr_decimal)
    qr_flat = qr_to_flat_dict(qr_decoded)

    full_text = pdf_text["fullText"]
    fields_out: list[dict] = []

    for row in schema:
        fn = row["fieldName"]
        sec = row["section"]
        qv, qc = qr_hint_matches_field(fn, sec, qr_flat)
        tv, evidence, tc = find_value_in_text(fn, sec, full_text)

        if qv and (tv is None or qc >= tc):
            fields_out.append(
                {
                    **row,
                    "status": "extracted",
                    "value": qv,
                    "source": "qr_inferred",
                    "confidence": qc,
                    "evidence": None,
                }
            )
        elif tv:
            fields_out.append(
                {
                    **row,
                    "status": "extracted",
                    "value": tv,
                    "source": "text_heuristic",
                    "confidence": tc,
                    "evidence": evidence,
                }
            )
        else:
            fields_out.append(
                {
                    **row,
                    "status": "not_found",
                    "value": None,
                    "source": None,
                    "confidence": 0.0,
                    "evidence": None,
                }
            )

    extracted = sum(1 for f in fields_out if f["status"] == "extracted")
    return {
        "pdfPath": _rel(pdf_path, ROOT),
        "schemaPath": _rel(schema_path, ROOT),
        "schemaRowCount": len(schema),
        "qr": {
            "decimalDigitCount": qr_decoded["decimalDigitCount"],
            "payloadByteLength": qr_decoded["payloadByteLength"],
            "segmentCount": qr_decoded["segmentCount"],
            "firstBinarySegmentIndex": qr_decoded["firstBinarySegmentIndex"],
            "segments": qr_decoded["segments"],
            "derivedFlat": qr_flat,
        },
        "pdfText": {
            "pageCount": pdf_text["pageCount"],
            "fullTextCharCount": pdf_text["fullTextCharCount"],
            "pages": [
                {
                    "page": p["page"],
                    "textCharCount": p["textCharCount"],
                    "lineCount": p["lineCount"],
                    "tableCount": p["tableCount"],
                }
                for p in pdf_text["pages"]
            ],
            "fullTextPreview": full_text[:4000] + ("…" if len(full_text) > 4000 else ""),
        },
        "fields": fields_out,
        "stats": {
            "extracted": extracted,
            "notFound": len(fields_out) - extracted,
            "fromQrInferred": sum(1 for f in fields_out if f.get("source") == "qr_inferred"),
            "fromTextHeuristic": sum(1 for f in fields_out if f.get("source") == "text_heuristic"),
        },
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True, help="Path to text-layer Shipping Bill PDF")
    ap.add_argument(
        "--schema",
        default=None,
        help="Override schema xlsx (default: Final_Schema/shipping_bill_fields.xlsx or fallback)",
    )
    ap.add_argument("--out", "-o", default=None, help="Write JSON result here")
    args = ap.parse_args()

    pdf_path = Path(args.pdf).resolve()
    if not pdf_path.is_file():
        print(f"PDF not found: {pdf_path}", file=sys.stderr)
        sys.exit(1)

    if args.schema:
        schema_path = Path(args.schema).resolve()
        if not schema_path.is_file():
            print(f"Schema not found: {schema_path}", file=sys.stderr)
            sys.exit(1)
    else:
        schema_path = resolve_schema_xlsx()

    try:
        result = run_extraction(pdf_path, schema_path)
    except Exception as e:
        print(e, file=sys.stderr)
        sys.exit(1)

    text = json.dumps(result, indent=2, ensure_ascii=False)
    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
        print(f"Wrote {args.out}")
    else:
        print(text)


if __name__ == "__main__":
    main()
